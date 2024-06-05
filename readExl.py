import openpyxl

def readIn(InPath):
    # Load the workbook and select the active worksheet
    wb = openpyxl.load_workbook(InPath)
    ws = wb.active

    data_2d = []
    for row in ws.iter_rows(min_row=1, max_row=ws.max_row, min_col=1, max_col=ws.max_column):
        row_data = [cell.value for cell in row]
        data_2d.append(row_data)
    
    data=''
    for b in range(len(data_2d[0])):
        temp = []
        for a in range(len(data_2d)):
            if data_2d[a][b] is not None and a>0:
                temp.append(str(data_2d[a][b]))

        if b==len(data_2d[0])-1:
            data+=','.join(temp)
        elif b>=0:
            data+=','.join(temp)+'||'

    return data


