import openpyxl
from openpyxl.styles import Border, Side, PatternFill, Font, Alignment


def template(col):
    data = col

    try:
        wb = openpyxl.Workbook()
        ws = wb.active

        # Write the array content horizontally
        for col, value in enumerate(data, start=1):
            c=ws.cell(row=1, column=col, value=value)

            thin = Side(border_style="thin", color="00000000")

            c.border = Border(top=thin, left=thin, right=thin, bottom=thin)
            c.fill = PatternFill("solid", fgColor="0000FFFF")
            c.font  = Font(b=True, color="00000000")
            c.alignment = Alignment(horizontal="center", vertical="center")

        # Save the workbook to a file
        wb.save('template.xlsx')
        # wb.close()
        
    except:
        return False
    

